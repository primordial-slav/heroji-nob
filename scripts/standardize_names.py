import csv
import re
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

def clean_name_part(name):
    """Clean and standardize a name part"""
    if not name or name == 'nan':
        return ''
    # Remove extra whitespace
    name = ' '.join(name.split())
    # Convert to title case if all caps (except for single letters/initials)
    if len(name) > 2 and name.isupper():
        name = name.title()
    return name.strip()

def parse_first_name(first_name_field):
    """Parse first name field which might contain father's name in CAPS"""
    if not first_name_field:
        return '', ''

    # Check if there's a pattern like "Vittorio SALVATORE" (first name + FATHER'S NAME)
    parts = first_name_field.split()
    if len(parts) >= 2:
        # Check if last part is all caps (likely father's name)
        if parts[-1].isupper() and len(parts[-1]) > 1:
            fathers_name = parts[-1].title()
            first_name = ' '.join(parts[:-1])
            return clean_name_part(first_name), clean_name_part(fathers_name)

    return clean_name_part(first_name_field), ''

def standardize_names():
    """Read all CSV files and create a standardized Excel file"""

    csv_files = [
        '01-data/processed/prva_licka_proleterska_soldiers.csv',
        '01-data/processed/prva_proleterska_soldiers.csv',
        '01-data/processed/ljubljanska_soldiers.csv'
    ]

    all_soldiers = []

    for csv_file in csv_files:
        file_path = Path(csv_file)
        if not file_path.exists():
            print(f"Warning: {csv_file} not found, skipping...")
            continue

        print(f"Processing {csv_file}...")

        # Determine source/brigade
        if 'prva_licka' in csv_file:
            source = 'Prva Lička Proleterska Brigada'
        elif 'prva_proleterska' in csv_file:
            source = 'Prva Proleterska Brigada'
        elif 'ljubljanska' in csv_file:
            source = 'Ljubljanska Brigada'
        else:
            source = 'Unknown Source'

        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                last_name = clean_name_part(row.get('last_name', ''))
                middle_name = clean_name_part(row.get('middle_name', ''))
                first_name_raw = row.get('first_name', '')

                # Parse first name (might contain father's name)
                first_name, fathers_name = parse_first_name(first_name_raw)

                # If we extracted a father's name and there's no middle name, use it as middle name
                if fathers_name and not middle_name:
                    middle_name = fathers_name

                additional_info = row.get('additional_info', '').strip()

                # Skip rows with no meaningful data
                if not last_name and not first_name:
                    continue

                all_soldiers.append({
                    'last_name': last_name,
                    'middle_name': middle_name,
                    'first_name': first_name,
                    'full_name': f"{last_name} {middle_name} {first_name}".replace('  ', ' ').strip(),
                    'additional_info': additional_info,
                    'source': source
                })

    # Sort by last name, then first name
    all_soldiers.sort(key=lambda x: (x['last_name'].lower(), x['first_name'].lower()))

    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Standardized Soldiers"

    # Define headers
    headers = ['Last Name', 'Middle Name', 'First Name', 'Full Name', 'Additional Information', 'Source/Brigade']

    # Style headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    for row_num, soldier in enumerate(all_soldiers, 2):
        ws.cell(row=row_num, column=1, value=soldier['last_name'])
        ws.cell(row=row_num, column=2, value=soldier['middle_name'])
        ws.cell(row=row_num, column=3, value=soldier['first_name'])
        ws.cell(row=row_num, column=4, value=soldier['full_name'])
        ws.cell(row=row_num, column=5, value=soldier['additional_info'])
        ws.cell(row=row_num, column=6, value=soldier['source'])

    # Adjust column widths
    ws.column_dimensions['A'].width = 20  # Last Name
    ws.column_dimensions['B'].width = 15  # Middle Name
    ws.column_dimensions['C'].width = 15  # First Name
    ws.column_dimensions['D'].width = 30  # Full Name
    ws.column_dimensions['E'].width = 50  # Additional Info
    ws.column_dimensions['F'].width = 25  # Source

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save file
    output_file = '01-data/processed/standardized_soldiers.xlsx'
    wb.save(output_file)

    print(f"\nSuccessfully created {output_file}")
    print(f"Total soldiers: {len(all_soldiers)}")
    print(f"Format: Last Name | Middle Name | First Name | Full Name | Additional Info | Source")

if __name__ == '__main__':
    standardize_names()
