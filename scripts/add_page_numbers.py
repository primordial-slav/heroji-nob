import csv
import re
import sys
from pathlib import Path

# Set UTF-8 encoding for output
sys.stdout.reconfigure(encoding='utf-8')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

def parse_extracted_text_with_pages(text_file_path):
    """
    Parse extracted text file and build a mapping of soldier entries to page numbers.
    Returns dict: {soldier_key: page_number}
    """
    print(f"Parsing {text_file_path}...")

    soldier_pages = {}
    current_page = 0
    in_soldier_section = False

    with open(text_file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()

            # Track page numbers - handle multiple formats
            # Format 1: "PAGE 557"
            if line.startswith('PAGE '):
                try:
                    current_page = int(line.split()[1])
                except:
                    pass
                continue

            # Format 2: "--- PAGE 9 of prva proleterska 1.pdf ---"
            if line.startswith('--- PAGE ') and ' of ' in line:
                try:
                    page_part = line.split('PAGE')[1].split('of')[0].strip()
                    current_page = int(page_part)
                except:
                    pass
                continue

            # Skip page separators
            if line == '=' * 50:
                continue

            # Detect soldier listing section (usually starts with uppercase names)
            # Pattern: LASTNAME Firstname, YYYY, Location...
            # or: LASTNAME FATHERSNAME, YYYY, Location...
            soldier_pattern = r'^([A-ZČĆŽŠĐ][A-ZČĆŽŠĐ]+)\s+(.+?)(?:,\s*(\d{4})|$)'
            match = re.match(soldier_pattern, line)

            if match and current_page > 0:
                last_name = match.group(1)
                rest = match.group(2)

                # Extract first part of rest (could be first name or father's name)
                first_parts = rest.split(',')[0].strip().split()
                if first_parts:
                    first_name = first_parts[0]

                    # Create a key for lookup
                    soldier_key = f"{last_name}_{first_name}".upper()

                    # Store the page number
                    if soldier_key not in soldier_pages:
                        soldier_pages[soldier_key] = current_page

    print(f"  Found {len(soldier_pages)} soldier entries with page numbers")
    return soldier_pages

def extract_date_of_birth(additional_info):
    """Extract date of birth from additional_info field."""
    if not additional_info or additional_info.strip() == '':
        return '', ''

    info = additional_info.strip()

    # Extract only the first part BEFORE any death mentions
    death_keywords = ['poginuo', 'padel', 'nestao', 'umrl']
    first_death_pos = len(info)
    for keyword in death_keywords:
        pos = info.lower().find(keyword)
        if pos != -1 and pos < first_death_pos:
            first_death_pos = pos

    search_text = info[:first_death_pos]

    # Pattern 1: Full date - DD. MM. YYYY
    match = re.match(r'^(\d{1,2}\.\s*\d{1,2}\.\s*\d{4})', search_text)
    if match:
        date_str = match.group(1)
        year_match = re.search(r'\d{4}', date_str)
        year = year_match.group(0) if year_match else ''
        return date_str.strip(), year

    # Pattern 2: Year with period - YYYY.
    match = re.match(r'^(\d{4})\.\s+\w', search_text)
    if match:
        year = match.group(1)
        return year, year

    # Pattern 3: Year with comma - YYYY,
    match = re.match(r'^(\d{4}),', search_text)
    if match:
        year = match.group(1)
        return year, year

    # Pattern 4: Year with space - YYYY
    match = re.match(r'^(\d{4})\s+[A-ZČĆŽŠĐ]', search_text)
    if match:
        year = match.group(1)
        return year, year

    return '', ''

def standardize_with_page_numbers():
    """Create Excel with page numbers for each soldier"""

    # Define source mappings
    sources = [
        {
            'csv': '01-data/processed/prva_licka_proleterska_soldiers.csv',
            'txt': '01-data/processed/extracted_text.txt',
            'brigade': 'Prva Lička Proleterska Brigada',
            'book': 'Prva Lička Proleterska NOV Brigada'
        },
        {
            'csv': '01-data/processed/prva_proleterska_soldiers.csv',
            'txt': '01-data/processed/prva_proleterska_extracted.txt',
            'brigade': 'Prva Proleterska Brigada',
            'book': 'Prva Proleterska Brigada - Zbornik'
        },
        {
            'csv': '01-data/processed/ljubljanska_soldiers.csv',
            'txt': '01-data/processed/ljubljanska_extracted.txt',
            'brigade': 'Ljubljanska Brigada',
            'book': 'Ljubljanska Brigada'
        }
    ]

    all_soldiers = []
    stats = {
        'total': 0,
        'with_page': 0,
        'without_page': 0,
        'with_full_date': 0,
        'with_year_only': 0,
        'no_date': 0
    }

    for source in sources:
        csv_file = Path(source['csv'])
        txt_file = Path(source['txt'])

        if not csv_file.exists():
            print(f"Warning: {csv_file} not found, skipping...")
            continue

        print(f"\nProcessing {source['brigade']}...")

        # Parse text file to get page numbers
        soldier_pages = {}
        if txt_file.exists():
            soldier_pages = parse_extracted_text_with_pages(txt_file)
        else:
            print(f"  Warning: {txt_file} not found, page numbers will be empty")

        # Read CSV and match soldiers to pages
        with open(csv_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                last_name = row.get('last_name', '').strip()
                middle_name = row.get('middle_name', '').strip()
                first_name = row.get('first_name', '').strip()
                additional_info = row.get('additional_info', '').strip()

                # Skip rows with no meaningful data
                if not last_name and not first_name:
                    continue

                # Extract date of birth
                dob, birth_year = extract_date_of_birth(additional_info)

                # Find page number
                soldier_key = f"{last_name}_{first_name}".upper()
                page_number = soldier_pages.get(soldier_key, '')

                # Update statistics
                stats['total'] += 1
                if page_number:
                    stats['with_page'] += 1
                else:
                    stats['without_page'] += 1

                if dob:
                    if '.' in dob and len(dob) > 4:
                        stats['with_full_date'] += 1
                    else:
                        stats['with_year_only'] += 1
                else:
                    stats['no_date'] += 1

                all_soldiers.append({
                    'last_name': last_name,
                    'middle_name': middle_name,
                    'first_name': first_name,
                    'full_name': f"{last_name} {middle_name} {first_name}".replace('  ', ' ').strip(),
                    'date_of_birth': dob,
                    'birth_year': birth_year,
                    'additional_info': additional_info,
                    'source': source['brigade'],
                    'book': source['book'],
                    'page': str(page_number) if page_number else ''
                })

    # Sort by last name, then first name
    all_soldiers.sort(key=lambda x: (x['last_name'].lower(), x['first_name'].lower()))

    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Soldiers with Sources"

    # Define headers
    headers = [
        'Last Name', 'Middle Name', 'First Name', 'Full Name',
        'Date of Birth', 'Birth Year', 'Source Book', 'Page',
        'Additional Information', 'Brigade'
    ]

    # Style headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    for row_num, soldier in enumerate(all_soldiers, 2):
        ws.cell(row=row_num, column=1, value=soldier['last_name'])
        ws.cell(row=row_num, column=2, value=soldier['middle_name'])
        ws.cell(row=row_num, column=3, value=soldier['first_name'])
        ws.cell(row=row_num, column=4, value=soldier['full_name'])
        ws.cell(row=row_num, column=5, value=soldier['date_of_birth'])
        ws.cell(row=row_num, column=6, value=soldier['birth_year'])
        ws.cell(row=row_num, column=7, value=soldier['book'])
        ws.cell(row=row_num, column=8, value=soldier['page'])
        ws.cell(row=row_num, column=9, value=soldier['additional_info'])
        ws.cell(row=row_num, column=10, value=soldier['source'])

    # Adjust column widths
    ws.column_dimensions['A'].width = 20  # Last Name
    ws.column_dimensions['B'].width = 15  # Middle Name
    ws.column_dimensions['C'].width = 15  # First Name
    ws.column_dimensions['D'].width = 30  # Full Name
    ws.column_dimensions['E'].width = 18  # Date of Birth
    ws.column_dimensions['F'].width = 12  # Birth Year
    ws.column_dimensions['G'].width = 35  # Source Book
    ws.column_dimensions['H'].width = 8   # Page
    ws.column_dimensions['I'].width = 50  # Additional Info
    ws.column_dimensions['J'].width = 25  # Brigade

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save file
    output_file = '01-data/processed/standardized_soldiers_with_pages.xlsx'
    wb.save(output_file)

    # Print statistics
    print(f"\nSuccessfully created {output_file}")
    print(f"\nStatistics:")
    print(f"  Total soldiers: {stats['total']}")
    print(f"  With page numbers: {stats['with_page']} ({stats['with_page']/stats['total']*100:.1f}%)")
    print(f"  Without page numbers: {stats['without_page']} ({stats['without_page']/stats['total']*100:.1f}%)")
    print(f"  With full date (DD.MM.YYYY): {stats['with_full_date']} ({stats['with_full_date']/stats['total']*100:.1f}%)")
    print(f"  With year only (YYYY): {stats['with_year_only']} ({stats['with_year_only']/stats['total']*100:.1f}%)")
    print(f"  No date found: {stats['no_date']} ({stats['no_date']/stats['total']*100:.1f}%)")

if __name__ == '__main__':
    standardize_with_page_numbers()
