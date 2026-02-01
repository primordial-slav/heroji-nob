import csv
import re
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

def extract_date_of_birth(additional_info):
    """
    Extract date of birth from additional_info field.
    Returns tuple: (date_of_birth_string, birth_year)

    Birth dates appear at the START of the additional_info field.
    Death dates (after 'poginuo', 'padel', 'nestao') are ignored.
    """
    if not additional_info or additional_info.strip() == '':
        return '', ''

    info = additional_info.strip()

    # Extract only the first part BEFORE any death mentions
    # This ensures we don't accidentally grab death dates
    death_keywords = ['poginuo', 'padel', 'nestao', 'umrl']

    # Find the earliest death keyword position
    first_death_pos = len(info)
    for keyword in death_keywords:
        pos = info.lower().find(keyword)
        if pos != -1 and pos < first_death_pos:
            first_death_pos = pos

    # Only look at the beginning part (before death mentions)
    search_text = info[:first_death_pos]

    # Pattern 1: Full date at start - DD. MM. YYYY, or D. M. YYYY,
    # Examples: "20. 11. 1927, Selevac", "1. 8. 1923, Zalogovac"
    match = re.match(r'^(\d{1,2}\.\s*\d{1,2}\.\s*\d{4})', search_text)
    if match:
        date_str = match.group(1)
        # Extract year
        year_match = re.search(r'\d{4}', date_str)
        year = year_match.group(0) if year_match else ''
        return date_str.strip(), year

    # Pattern 2: Year with period - YYYY. Location
    # Examples: "1913. Tinjan", "1920. Trbovlje"
    match = re.match(r'^(\d{4})\.\s+\w', search_text)
    if match:
        year = match.group(1)
        return year, year

    # Pattern 3: Year with comma - YYYY, Location (most common: 61.6%)
    # Examples: "1906, Volčja jama", "1915, Slivovo"
    match = re.match(r'^(\d{4}),', search_text)
    if match:
        year = match.group(1)
        return year, year

    # Pattern 4: Year with space - YYYY Location
    # Examples: "1923 Kožljevec", "1926 Velika"
    match = re.match(r'^(\d{4})\s+[A-ZČĆŽŠĐ]', search_text)
    if match:
        year = match.group(1)
        return year, year

    # No clear birth date found
    return '', ''

def standardize_with_dates():
    """Read all CSV files and create Excel with date of birth column"""

    csv_files = [
        '01-data/processed/prva_licka_proleterska_soldiers.csv',
        '01-data/processed/prva_proleterska_soldiers.csv',
        '01-data/processed/ljubljanska_soldiers.csv'
    ]

    all_soldiers = []
    stats = {
        'total': 0,
        'with_full_date': 0,
        'with_year_only': 0,
        'no_date': 0
    }

    for csv_file in csv_files:
        file_path = Path(csv_file)
        if not file_path.exists():
            print(f"Warning: {csv_file} not found, skipping...")
            continue

        print(f"Processing {csv_file}...")

        # Determine source/brigade
        if 'prva_licka' in csv_file:
            source = 'Prva Lička Proleterska Brigada'
        elif 'prva_proleterska' in csv_file:
            source = 'Prva Proleterska Brigada'
        elif 'ljubljanska' in csv_file:
            source = 'Ljubljanska Brigada'
        else:
            source = 'Unknown Source'

        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                last_name = row.get('last_name', '').strip()
                middle_name = row.get('middle_name', '').strip()
                first_name = row.get('first_name', '').strip()
                additional_info = row.get('additional_info', '').strip()

                # Skip rows with no meaningful data
                if not last_name and not first_name:
                    continue

                # Extract date of birth
                dob, birth_year = extract_date_of_birth(additional_info)

                # Update statistics
                stats['total'] += 1
                if dob:
                    if '.' in dob and len(dob) > 4:
                        stats['with_full_date'] += 1
                    else:
                        stats['with_year_only'] += 1
                else:
                    stats['no_date'] += 1

                all_soldiers.append({
                    'last_name': last_name,
                    'middle_name': middle_name,
                    'first_name': first_name,
                    'full_name': f"{last_name} {middle_name} {first_name}".replace('  ', ' ').strip(),
                    'date_of_birth': dob,
                    'birth_year': birth_year,
                    'additional_info': additional_info,
                    'source': source
                })

    # Sort by last name, then first name
    all_soldiers.sort(key=lambda x: (x['last_name'].lower(), x['first_name'].lower()))

    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Soldiers with DOB"

    # Define headers
    headers = ['Last Name', 'Middle Name', 'First Name', 'Full Name', 'Date of Birth', 'Birth Year', 'Additional Information', 'Source/Brigade']

    # Style headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    for row_num, soldier in enumerate(all_soldiers, 2):
        ws.cell(row=row_num, column=1, value=soldier['last_name'])
        ws.cell(row=row_num, column=2, value=soldier['middle_name'])
        ws.cell(row=row_num, column=3, value=soldier['first_name'])
        ws.cell(row=row_num, column=4, value=soldier['full_name'])
        ws.cell(row=row_num, column=5, value=soldier['date_of_birth'])
        ws.cell(row=row_num, column=6, value=soldier['birth_year'])
        ws.cell(row=row_num, column=7, value=soldier['additional_info'])
        ws.cell(row=row_num, column=8, value=soldier['source'])

    # Adjust column widths
    ws.column_dimensions['A'].width = 20  # Last Name
    ws.column_dimensions['B'].width = 15  # Middle Name
    ws.column_dimensions['C'].width = 15  # First Name
    ws.column_dimensions['D'].width = 30  # Full Name
    ws.column_dimensions['E'].width = 18  # Date of Birth
    ws.column_dimensions['F'].width = 12  # Birth Year
    ws.column_dimensions['G'].width = 50  # Additional Info
    ws.column_dimensions['H'].width = 25  # Source

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save file
    output_file = '01-data/processed/standardized_soldiers.xlsx'
    wb.save(output_file)

    # Print statistics
    print(f"\nSuccessfully created {output_file}")
    print(f"\nStatistics:")
    print(f"  Total soldiers: {stats['total']}")
    print(f"  With full date (DD.MM.YYYY): {stats['with_full_date']} ({stats['with_full_date']/stats['total']*100:.1f}%)")
    print(f"  With year only (YYYY): {stats['with_year_only']} ({stats['with_year_only']/stats['total']*100:.1f}%)")
    print(f"  No date found: {stats['no_date']} ({stats['no_date']/stats['total']*100:.1f}%)")
    print(f"  Total with birth info: {stats['with_full_date'] + stats['with_year_only']} ({(stats['with_full_date'] + stats['with_year_only'])/stats['total']*100:.1f}%)")

if __name__ == '__main__':
    standardize_with_dates()
